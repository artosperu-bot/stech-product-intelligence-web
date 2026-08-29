from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook

_BARCODE_LENGTHS = {8, 12, 13, 14}
_HEADER_PRIORITIES = (
    ('sku del vendedor', 100),
    ('part number', 98),
    ('partnumber', 98),
    ('mpn', 98),
    ('modelo', 90),
    ('codigo de barras', 80),
    ('código de barras', 80),
    ('nombre', 50),
)


@dataclass(frozen=True)
class IdentityCandidate:
    value: str
    kind: str
    field_name: str
    row: int
    priority: int


@dataclass
class CanonicalIdentity:
    brand: str = ''
    manufacturer_part_number: str = ''
    commercial_model: str = ''
    ean_upc_gtin: list[str] = field(default_factory=list)
    variant: str = ''
    color: str = ''
    capacity: str = ''
    region: str = ''
    aliases: list[str] = field(default_factory=list)
    confidence: int = 0
    sources: list[dict] = field(default_factory=list)


def _text(value: Any) -> str:
    return str(value or '').strip()


def _header_key(value: Any) -> str:
    return ' '.join(_text(value).casefold().replace('_', ' ').split())


def classify_identifier(value: str, field_name: str = '') -> str:
    text = _text(value)
    if not text:
        return 'UNKNOWN'
    compact = re.sub(r'\s+', '', text)
    if compact.isdigit() and len(compact) in _BARCODE_LENGTHS:
        return 'EAN_UPC_GTIN'
    header = _header_key(field_name)
    if 'nombre' in header and (' ' in text or len(text) > 40):
        return 'TEXT_ALIAS'
    if (
        re.fullmatch(r'[A-Za-z0-9._/-]+', compact)
        and any(c.isalpha() for c in compact)
        and any(c.isdigit() for c in compact)
    ):
        return 'PART_NUMBER'
    if 'modelo' in header:
        return 'MODEL'
    if 'sku' in header:
        return 'SELLER_SKU'
    if ' ' in text:
        return 'TEXT_ALIAS'
    return 'UNKNOWN'


def _priority_for_header(header: str) -> int:
    normalized = _header_key(header)
    for needle, priority in _HEADER_PRIORITIES:
        if needle in normalized:
            return priority
    return 20


def _find_header_row(ws) -> tuple[int, dict[int, str]]:
    best_row = 0
    best_headers: dict[int, str] = {}
    best_score = 0
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=min(ws.max_row, 30)),
        start=1,
    ):
        headers: dict[int, str] = {}
        score = 0
        for cell in row:
            text = _text(cell.value)
            if not text:
                continue
            priority = _priority_for_header(text)
            if priority > 20:
                headers[cell.column] = text
                score += 1
        if score > best_score:
            best_score = score
            best_row = row_idx
            best_headers = headers
    if best_score == 0:
        raise ValueError('TEMPLATE_HEADER_NOT_FOUND')
    return best_row, best_headers


def _evidence_candidates(wb) -> list[IdentityCandidate]:
    if 'IA_EVIDENCIA' not in wb.sheetnames:
        return []
    ws = wb['IA_EVIDENCIA']
    if ws.max_row < 2:
        return []
    headers = {_header_key(cell.value): cell.column for cell in ws[1] if _text(cell.value)}
    pn_col = headers.get('part number') or headers.get('partnumber') or headers.get('mpn')
    status_col = headers.get('estado') or headers.get('status')
    confidence_col = headers.get('confianza') or headers.get('confidence')
    if not pn_col:
        return []
    result: list[IdentityCandidate] = []
    seen: set[str] = set()
    for row_idx in range(2, ws.max_row + 1):
        value = _text(ws.cell(row_idx, pn_col).value)
        if not value or value in seen:
            continue
        status = _text(ws.cell(row_idx, status_col).value).upper() if status_col else ''
        if status_col and status not in {'CONFIRMADO', 'CONFIRMED'}:
            continue
        try:
            confidence = int(float(ws.cell(row_idx, confidence_col).value or 0)) if confidence_col else 100
        except Exception:
            confidence = 0
        if confidence_col and confidence < 80:
            continue
        if classify_identifier(value, 'Part Number') != 'PART_NUMBER':
            continue
        seen.add(value)
        result.append(IdentityCandidate(value, 'PART_NUMBER', 'IA_EVIDENCIA.Part Number', row_idx, 140))
    return result


def extract_identity_candidates(path: Path) -> list[IdentityCandidate]:
    wb = load_workbook(path, data_only=False, read_only=True)
    try:
        ws = wb['Subir plantilla'] if 'Subir plantilla' in wb.sheetnames else wb[wb.sheetnames[0]]
        header_row, headers = _find_header_row(ws)
        candidates: list[IdentityCandidate] = []
        seen: set[tuple[str, str, int]] = set()
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=header_row + 1),
            start=header_row + 1,
        ):
            for column, field_name in headers.items():
                value = _text(row[column - 1].value)
                if not value:
                    continue
                kind = classify_identifier(value, field_name)
                if kind == 'UNKNOWN':
                    continue
                key = (value, field_name, row_idx)
                if key in seen:
                    continue
                seen.add(key)
                priority = _priority_for_header(field_name)
                if kind == 'EAN_UPC_GTIN':
                    priority = min(priority, 80)
                candidates.append(IdentityCandidate(value, kind, field_name, row_idx, priority))
        candidates.extend(_evidence_candidates(wb))
        candidates.sort(key=lambda c: (-c.priority, c.row, c.field_name, c.value))
        return candidates
    finally:
        wb.close()


def choose_research_identifier(
    manual_identifier: str | None,
    candidates: list[IdentityCandidate],
) -> tuple[str, str, str]:
    manual = _text(manual_identifier)
    if manual:
        return 'manual', manual, classify_identifier(manual)
    usable = [
        candidate
        for candidate in candidates
        if candidate.kind in {'PART_NUMBER', 'MODEL', 'EAN_UPC_GTIN', 'SELLER_SKU'}
    ]
    if not usable:
        raise ValueError('IDENTITY_CANDIDATE_NOT_FOUND')
    chosen = sorted(usable, key=lambda c: (-c.priority, c.row))[0]
    return 'auto', chosen.value, chosen.kind


def _first(mapping: dict, keys: tuple[str, ...]) -> str:
    folded = {str(key).casefold(): value for key, value in mapping.items()}
    for key in keys:
        value = _text(folded.get(key.casefold()))
        if value:
            return value
    return ''


def _codes(mapping: dict) -> list[str]:
    result: list[str] = []
    folded = {str(key).casefold(): value for key, value in mapping.items()}
    for key in ('ean_upc_gtin', 'ean', 'upc', 'gtin', 'codigo_barras', 'código_barras'):
        value = folded.get(key.casefold())
        items = value if isinstance(value, list) else [value]
        for item in items:
            text = _text(item)
            if text and text not in result:
                result.append(text)
    return result


def canonical_identity_from_raw(raw: dict, fallback_identifier: str = '') -> CanonicalIdentity:
    raw = raw if isinstance(raw, dict) else {}
    product = raw.get('identity') or raw.get('producto') or raw.get('product') or raw
    product = product if isinstance(product, dict) else {}
    mpn = _first(
        product,
        (
            'manufacturer_part_number',
            'part_number',
            'partnumber',
            'mpn',
            'sku_fabricante',
            'manufacturer_sku',
        ),
    )
    fallback = _text(fallback_identifier)
    if not mpn and fallback and classify_identifier(fallback) == 'PART_NUMBER':
        mpn = fallback
    model = _first(product, ('commercial_model', 'modelo_comercial', 'modelo', 'model'))
    brand = _first(product, ('brand', 'marca', 'manufacturer', 'fabricante'))
    aliases = product.get('aliases') or product.get('alias') or []
    if isinstance(aliases, str):
        aliases = [aliases]
    sources = product.get('sources') or product.get('fuentes') or []
    if not isinstance(sources, list):
        sources = [sources]
    try:
        confidence = int(float(product.get('confidence', product.get('confianza', 0)) or 0))
    except Exception:
        confidence = 0
    return CanonicalIdentity(
        brand=brand,
        manufacturer_part_number=mpn,
        commercial_model=model,
        ean_upc_gtin=_codes(product),
        variant=_first(product, ('variant', 'variante')),
        color=_first(product, ('color',)),
        capacity=_first(product, ('capacity', 'capacidad')),
        region=_first(product, ('region', 'región')),
        aliases=[_text(item) for item in aliases if _text(item)],
        confidence=confidence,
        sources=[item for item in sources if isinstance(item, dict)],
    )


def enrich_identity_from_workbook_evidence(path: Path, identity: CanonicalIdentity) -> CanonicalIdentity:
    wb = load_workbook(path, data_only=False, read_only=True)
    try:
        if 'IA_EVIDENCIA' not in wb.sheetnames:
            return identity
        ws = wb['IA_EVIDENCIA']
        if ws.max_row < 2:
            return identity
        headers = {_header_key(cell.value): cell.column for cell in ws[1] if _text(cell.value)}
        pn_col = headers.get('part number') or headers.get('partnumber') or headers.get('mpn')
        field_col = headers.get('campo') or headers.get('campo plantilla')
        value_col = headers.get('valor') or headers.get('valor propuesto') or headers.get('valor escrito')
        status_col = headers.get('estado') or headers.get('status')
        confidence_col = headers.get('confianza') or headers.get('confidence')
        source_col = headers.get('fuente') or headers.get('source url')
        source_type_col = headers.get('tipo fuente') or headers.get('source type')
        if not field_col or not value_col:
            return identity

        mpn = identity.manufacturer_part_number
        brand = identity.brand
        model = identity.commercial_model
        codes = list(identity.ean_upc_gtin)
        sources = list(identity.sources)
        confidence = identity.confidence

        for row_idx in range(2, ws.max_row + 1):
            status = _text(ws.cell(row_idx, status_col).value).upper() if status_col else 'CONFIRMED'
            if status not in {'CONFIRMADO', 'CONFIRMED'}:
                continue
            try:
                row_confidence = int(float(ws.cell(row_idx, confidence_col).value or 0)) if confidence_col else 100
            except Exception:
                row_confidence = 0
            if row_confidence < 80:
                continue
            row_pn = _text(ws.cell(row_idx, pn_col).value) if pn_col else ''
            if mpn and row_pn and row_pn.casefold() != mpn.casefold():
                continue
            if not mpn and row_pn and classify_identifier(row_pn, 'Part Number') == 'PART_NUMBER':
                mpn = row_pn
            field_name = _header_key(ws.cell(row_idx, field_col).value)
            value = _text(ws.cell(row_idx, value_col).value)
            if not value or value == '89':
                continue
            if not brand and field_name.startswith('marca'):
                brand = value
            elif not model and field_name.startswith('modelo'):
                model = value
            elif ('código de barras' in field_name or 'codigo de barras' in field_name) and value not in codes:
                if classify_identifier(value, 'Código de barras') == 'EAN_UPC_GTIN':
                    codes.append(value)
            source_url = _text(ws.cell(row_idx, source_col).value) if source_col else ''
            source_type = _text(ws.cell(row_idx, source_type_col).value) if source_type_col else 'WORKBOOK_EVIDENCE'
            if source_url and not any(_text(item.get('url') or item.get('source_url')) == source_url for item in sources if isinstance(item, dict)):
                sources.append({'url': source_url, 'source_type': source_type or 'WORKBOOK_EVIDENCE'})
            confidence = max(confidence, row_confidence)

        identity.brand = brand
        identity.manufacturer_part_number = mpn
        identity.commercial_model = model
        identity.ean_upc_gtin = codes
        identity.sources = sources
        identity.confidence = confidence
        return identity
    finally:
        wb.close()

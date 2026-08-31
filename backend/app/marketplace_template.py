from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
import re
import unicodedata
from typing import Any, Iterable

from openpyxl import load_workbook

_BARCODE_LENGTHS = {8, 12, 13, 14}


def _text(value: Any) -> str:
    return str(value or '').strip()


def _norm(value: Any) -> str:
    text = unicodedata.normalize('NFKD', _text(value).casefold())
    text = ''.join(ch for ch in text if not unicodedata.combining(ch))
    return ' '.join(text.replace('_', ' ').split())


def _compact_norm(value: Any) -> str:
    return re.sub(r'[^a-z0-9]+', '', _norm(value))


def classify_identifier(value: Any, field_name: str = '') -> str:
    text = _text(value)
    if not text:
        return 'UNKNOWN'
    compact = re.sub(r'\s+', '', text)
    if compact.isdigit() and len(compact) in _BARCODE_LENGTHS:
        return 'EAN_UPC_GTIN'
    header = _norm(field_name)
    if 'nombre' in header and (' ' in text or len(text) > 40):
        return 'TEXT_ALIAS'
    if (
        re.fullmatch(r'[A-Za-z0-9._/-]+', compact)
        and any(ch.isalpha() for ch in compact)
        and any(ch.isdigit() for ch in compact)
    ):
        return 'PART_NUMBER'
    if 'modelo' in header or header == 'model':
        return 'MODEL'
    if 'sku' in header:
        return 'SELLER_SKU'
    if ' ' in text:
        return 'TEXT_ALIAS'
    return 'UNKNOWN'


@dataclass(frozen=True)
class TemplateField:
    column: int
    column_letter: str
    label: str
    code: str
    group: str = ''
    instruction: str = ''
    example_value: str = ''
    requirements: dict[str, str] = field(default_factory=dict)
    valid_values: tuple[str, ...] = ()

    def requirement_for(self, category: str = '') -> str:
        if category and category in self.requirements:
            return self.requirements[category]
        if '*' in self.requirements:
            return self.requirements['*']
        return 'OPTIONAL'


@dataclass(frozen=True)
class ProductSlot:
    row: int
    identifier: str
    identifier_type: str
    category: str
    existing_values: dict[str, str]
    identity_source: str


@dataclass
class MarketplaceTemplateProfile:
    marketplace: str
    sheet_name: str
    header_rows: tuple[int, ...]
    data_start_row: int
    fields: list[TemplateField]
    products: list[ProductSlot]
    category_options: tuple[str, ...] = ()
    warnings: list[str] = field(default_factory=list)

    @property
    def product_count(self) -> int:
        return len(self.products)

    def field_by_code(self, code: str) -> TemplateField | None:
        target = _compact_norm(code)
        for item in self.fields:
            if _compact_norm(item.code) == target:
                return item
        return None

    def field_by_label(self, *labels: str) -> TemplateField | None:
        targets = {_compact_norm(item) for item in labels if _text(item)}
        for item in self.fields:
            base = re.sub(r'\s*#\s*\d+\s*$', '', item.label)
            if _compact_norm(item.label) in targets or _compact_norm(base) in targets:
                return item
        return None


_OPTION_ALIASES = {
    'paisdeproduccion': 'productioncountry',
    'condiciondelproducto': 'conditiontype',
    'garantiadelproducto': 'sellerwarranty',
}


def _column_letter(index: int) -> str:
    result = ''
    n = index
    while n:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result


def _row_values(ws, row_idx: int, max_col: int) -> tuple[Any, ...]:
    return next(ws.iter_rows(min_row=row_idx, max_row=row_idx, max_col=max_col, values_only=True))


def _find_falabella_header(ws, scan_rows: int = 15, scan_cols: int = 256) -> tuple[int, tuple[Any, ...]]:
    best_row = 0
    best_values: tuple[Any, ...] = ()
    best_score = 0
    for row_idx, values in enumerate(
        ws.iter_rows(min_row=1, max_row=scan_rows, max_col=scan_cols, values_only=True),
        start=1,
    ):
        score = sum(1 for value in values if re.search(r'#\s*\d+\b', _text(value)))
        if score > best_score:
            best_row, best_values, best_score = row_idx, values, score
    if best_score < 3:
        raise ValueError('FALABELLA_HEADER_NOT_FOUND')
    return best_row, best_values


def _extract_instruction_example(instruction: str) -> str:
    text = _text(instruction)
    if not text:
        return ''
    match = re.search(r'(?im)^\s*-\s*Value\s*:\s*(.+?)\s*$', text)
    return _text(match.group(1)) if match else ''


def _load_column_lists(ws, max_rows: int = 20000) -> dict[str, tuple[str, ...]]:
    rows = list(ws.iter_rows(min_row=1, max_row=max_rows, values_only=True))
    if not rows:
        return {}
    headers = rows[0]
    result: dict[str, tuple[str, ...]] = {}
    for col_idx, header in enumerate(headers):
        key = _compact_norm(header)
        if not key:
            continue
        values: list[str] = []
        seen: set[str] = set()
        for row in rows[1:]:
            if col_idx >= len(row):
                continue
            value = _text(row[col_idx])
            if not value or value in seen:
                continue
            seen.add(value)
            values.append(value)
        result[key] = tuple(values)
    return result


def _is_placeholder(field: TemplateField, value: Any) -> bool:
    text = _text(value)
    if not text:
        return True
    if field.example_value and _norm(text) == _norm(field.example_value):
        return True
    normalized = _norm(text)
    if normalized in {'...', 'esto es un parrafo'}:
        return True
    return False


def is_placeholder_value(field: TemplateField, value: Any) -> bool:
    return _is_placeholder(field, value)


def _value_for(slot_values: dict[str, str], field: TemplateField | None) -> str:
    if field is None:
        return ''
    return _text(slot_values.get(field.code))


def _pick_falabella_identifier(profile: MarketplaceTemplateProfile, values: dict[str, str]) -> tuple[str, str, str] | None:
    sku = profile.field_by_label('SKU del vendedor')
    model = profile.field_by_label('Modelo')
    barcode = profile.field_by_label('Código de barras', 'Codigo de barras')
    name = profile.field_by_label('Nombre')

    candidates: list[tuple[int, str, str, str]] = []
    if sku:
        value = _value_for(values, sku)
        if value and not _is_placeholder(sku, value):
            kind = classify_identifier(value, sku.label)
            if kind != 'UNKNOWN':
                candidates.append((120, value, 'PART_NUMBER' if kind == 'PART_NUMBER' else kind, sku.code))
    if model:
        value = _value_for(values, model)
        if value and not _is_placeholder(model, value):
            kind = classify_identifier(value, model.label)
            candidates.append((110 if kind == 'PART_NUMBER' else 95, value, kind, model.code))
    if barcode:
        value = _value_for(values, barcode)
        if value and not _is_placeholder(barcode, value):
            kind = classify_identifier(value, barcode.label)
            if kind == 'EAN_UPC_GTIN':
                candidates.append((85, value, kind, barcode.code))
    if name:
        value = _value_for(values, name)
        if value and not _is_placeholder(name, value):
            candidates.append((40, value, 'TEXT_ALIAS', name.code))
    if not candidates:
        return None
    _, value, kind, source = sorted(candidates, key=lambda item: -item[0])[0]
    return value, kind, source


def _pick_ripley_identifier(profile: MarketplaceTemplateProfile, values: dict[str, str]) -> tuple[str, str, str] | None:
    model = profile.field_by_code('modelo')
    barcode = profile.field_by_code('ean')
    seller_sku = profile.field_by_code('sku_seller')
    name = profile.field_by_code('nombre')
    description = profile.field_by_code('descripcion')
    candidates: list[tuple[int, str, str, str]] = []

    if model:
        value = _value_for(values, model)
        if value and not _is_placeholder(model, value):
            kind = classify_identifier(value, model.label)
            candidates.append((120 if kind == 'PART_NUMBER' else 100, value, kind, model.code))
    if description:
        value = _value_for(values, description)
        match = re.search(r'(?i)\bPART\s*NUMBER\s*[:\-]\s*([A-Za-z0-9._/-]+)', value)
        if match:
            pn = match.group(1).strip()
            candidates.append((115, pn, 'PART_NUMBER', description.code))
    if barcode:
        value = _value_for(values, barcode)
        if value and not _is_placeholder(barcode, value) and classify_identifier(value, barcode.label) == 'EAN_UPC_GTIN':
            candidates.append((90, value, 'EAN_UPC_GTIN', barcode.code))
    if seller_sku:
        value = _value_for(values, seller_sku)
        if value and not _is_placeholder(seller_sku, value):
            candidates.append((60, value, 'SELLER_SKU', seller_sku.code))
    if name:
        value = _value_for(values, name)
        if value and not _is_placeholder(name, value):
            candidates.append((40, value, 'TEXT_ALIAS', name.code))
    if not candidates:
        return None
    _, value, kind, source = sorted(candidates, key=lambda item: -item[0])[0]
    return value, kind, source


def _falabella_profile(wb) -> MarketplaceTemplateProfile:
    ws = wb['Subir plantilla']
    header_row, header_values = _find_falabella_header(ws)
    max_col = max(idx for idx, value in enumerate(header_values, start=1) if _text(value))
    group_values = _row_values(ws, max(1, header_row - 3), max_col) if header_row >= 4 else tuple('' for _ in range(max_col))
    instruction_values = _row_values(ws, max(1, header_row - 2), max_col) if header_row >= 3 else tuple('' for _ in range(max_col))
    requirement_values = _row_values(ws, max(1, header_row - 1), max_col) if header_row >= 2 else tuple('' for _ in range(max_col))

    option_lists: dict[str, tuple[str, ...]] = {}
    if 'Opciones' in wb.sheetnames:
        option_lists = _load_column_lists(wb['Opciones'])

    category_options: list[str] = []
    if 'Categorías' in wb.sheetnames:
        cws = wb['Categorías']
        for row in cws.iter_rows(min_row=2, max_col=1, values_only=True):
            value = _text(row[0] if row else '')
            if value:
                category_options.append(value)

    fields: list[TemplateField] = []
    for col_idx in range(1, max_col + 1):
        label = _text(header_values[col_idx - 1])
        if not label:
            continue
        base_label = re.sub(r'\s*#\s*\d+\s*$', '', label).strip()
        code = base_label
        instruction = _text(instruction_values[col_idx - 1] if col_idx - 1 < len(instruction_values) else '')
        requirement_raw = _norm(requirement_values[col_idx - 1] if col_idx - 1 < len(requirement_values) else '')
        requirement = 'OPTIONAL' if 'optional' in requirement_raw else 'REQUIRED'
        example = _extract_instruction_example(instruction)
        if _compact_norm(base_label) in {'categoriaprimaria'} and category_options:
            valid_values = tuple(category_options)
        else:
            option_key = _compact_norm(base_label)
            option_key = _OPTION_ALIASES.get(option_key, option_key)
            valid_values = option_lists.get(option_key, ())
        fields.append(TemplateField(
            column=col_idx,
            column_letter=_column_letter(col_idx),
            label=label,
            code=code,
            group=_text(group_values[col_idx - 1] if col_idx - 1 < len(group_values) else ''),
            instruction=instruction,
            example_value=example,
            requirements={'*': requirement},
            valid_values=tuple(valid_values),
        ))

    profile = MarketplaceTemplateProfile(
        marketplace='falabella',
        sheet_name='Subir plantilla',
        header_rows=(header_row - 3, header_row - 2, header_row - 1, header_row),
        data_start_row=header_row + 1,
        fields=fields,
        products=[],
        category_options=tuple(category_options),
    )
    category_field = profile.field_by_label('Categoría primaria')
    category_default = category_options[0] if len(category_options) == 1 else ''
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=profile.data_start_row, max_col=max_col, values_only=True),
        start=profile.data_start_row,
    ):
        existing = {field.code: _text(row[field.column - 1] if field.column - 1 < len(row) else '') for field in fields}
        picked = _pick_falabella_identifier(profile, existing)
        if not picked:
            continue
        identifier, identifier_type, source = picked
        category = _value_for(existing, category_field) or category_default
        profile.products.append(ProductSlot(row_idx, identifier, identifier_type, category, existing, source))
    return profile


def _ripley_profile(wb) -> MarketplaceTemplateProfile:
    ws = wb['Data']
    header1 = _row_values(ws, 1, 256)
    header2 = _row_values(ws, 2, 256)
    max_col = max(
        [idx for idx, value in enumerate(header1, start=1) if _text(value)]
        + [idx for idx, value in enumerate(header2, start=1) if _text(value)]
    )

    columns_meta: dict[str, dict[str, Any]] = {}
    category_columns: dict[str, int] = {}
    if 'Columns' in wb.sheetnames:
        cws = wb['Columns']
        cheaders = _row_values(cws, 1, 256)
        for idx, value in enumerate(cheaders, start=1):
            if idx >= 5 and _text(value):
                category_columns[_text(value)] = idx
        for row in cws.iter_rows(min_row=2, max_col=max(4, max(category_columns.values(), default=4)), values_only=True):
            code = _text(row[0] if len(row) > 0 else '')
            if not code:
                continue
            meta = {
                'label': _text(row[1] if len(row) > 1 else ''),
                'instruction': _text(row[2] if len(row) > 2 else ''),
                'example': _text(row[3] if len(row) > 3 else ''),
                'requirements': {},
            }
            for category, col_idx in category_columns.items():
                status = _text(row[col_idx - 1] if col_idx - 1 < len(row) else '').upper()
                if status:
                    meta['requirements'][category] = status
            columns_meta[_compact_norm(code)] = meta

    reference_lists: dict[str, tuple[str, ...]] = {}
    if 'ReferenceData' in wb.sheetnames:
        reference_lists = _load_column_lists(wb['ReferenceData'])

    fields: list[TemplateField] = []
    for col_idx in range(1, max_col + 1):
        label = _text(header1[col_idx - 1] if col_idx - 1 < len(header1) else '')
        code = _text(header2[col_idx - 1] if col_idx - 1 < len(header2) else '') or label
        if not code:
            continue
        meta = columns_meta.get(_compact_norm(code), {})
        fields.append(TemplateField(
            column=col_idx,
            column_letter=_column_letter(col_idx),
            label=label or _text(meta.get('label')) or code,
            code=code,
            instruction=_text(meta.get('instruction')),
            example_value=_text(meta.get('example')),
            requirements=dict(meta.get('requirements') or {}),
            valid_values=reference_lists.get(_compact_norm(code), ()),
        ))

    profile = MarketplaceTemplateProfile(
        marketplace='ripley',
        sheet_name='Data',
        header_rows=(1, 2),
        data_start_row=3,
        fields=fields,
        products=[],
        category_options=tuple(category_columns),
    )
    category_field = profile.field_by_code('categoria')
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=3, max_col=max_col, values_only=True),
        start=3,
    ):
        existing = {field.code: _text(row[field.column - 1] if field.column - 1 < len(row) else '') for field in fields}
        picked = _pick_ripley_identifier(profile, existing)
        if not picked:
            continue
        identifier, identifier_type, source = picked
        category = _value_for(existing, category_field)
        profile.products.append(ProductSlot(row_idx, identifier, identifier_type, category, existing, source))
    return profile


def analyze_marketplace_template(path: Path) -> MarketplaceTemplateProfile:
    path = Path(path)
    wb = load_workbook(path, data_only=False, read_only=True)
    try:
        names = set(wb.sheetnames)
        if 'Subir plantilla' in names:
            return _falabella_profile(wb)
        if {'Data', 'Columns'}.issubset(names):
            return _ripley_profile(wb)
        raise ValueError('UNSUPPORTED_MARKETPLACE_TEMPLATE')
    finally:
        wb.close()


def choose_manual_slot(profile: MarketplaceTemplateProfile, identifier: str) -> ProductSlot | None:
    target = _norm(identifier)
    if not target:
        return None
    for slot in profile.products:
        if _norm(slot.identifier) == target:
            return slot
        if any(_norm(value) == target for value in slot.existing_values.values() if _text(value)):
            return slot
    if len(profile.products) == 1:
        slot = profile.products[0]
        return ProductSlot(slot.row, _text(identifier), classify_identifier(identifier), slot.category, slot.existing_values, 'manual')
    return None


def iter_required_fields(profile: MarketplaceTemplateProfile, slot: ProductSlot) -> Iterable[TemplateField]:
    for field in profile.fields:
        if field.requirement_for(slot.category).upper() in {'REQUIRED', 'MANDATORY'}:
            yield field

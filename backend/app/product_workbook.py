from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Font

from .product_evidence import MasterSpecification
from .product_identity import CanonicalIdentity

_REQUIRED_HEADERS = {
    'brand': ('marca #26', 'marca'),
    'sku': ('sku del vendedor #29', 'sku del vendedor'),
    'model': ('modelo #32', 'modelo'),
    'barcode': ('código de barras #56', 'codigo de barras #56', 'código de barras', 'codigo de barras'),
}

_EVIDENCE_HEADERS = [
    'Part Number', 'Campo plantilla', 'Valor propuesto', 'Valor escrito', 'Estado',
    'Confianza', 'Fuente', 'Tipo fuente', 'Página PDF', 'Evidencia', 'Observación',
]
_MASTER_HEADERS = [
    'Part Number', 'Categoría técnica', 'Especificación', 'Valor', 'Unidad', 'Estado',
    'Confianza', 'Fuente principal', 'Tipo fuente', 'Página PDF', 'Observación',
]


@dataclass
class ProductWorkbookQA:
    ok: bool
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    target_row: int | None = None


def _text(value: Any) -> str:
    return str(value or '').strip()


def _norm(value: Any) -> str:
    return ' '.join(_text(value).casefold().replace('_', ' ').split())


def _find_header_row(ws) -> tuple[int, dict[str, int]]:
    best: tuple[int, int, dict[str, int]] | None = None
    for row_idx in range(1, min(ws.max_row, 30) + 1):
        found: dict[str, int] = {}
        for col_idx in range(1, ws.max_column + 1):
            header = _norm(ws.cell(row_idx, col_idx).value)
            if not header:
                continue
            for key, aliases in _REQUIRED_HEADERS.items():
                if any(header == _norm(alias) for alias in aliases):
                    found[key] = col_idx
        score = len(found)
        if best is None or score > best[0]:
            best = (score, row_idx, found)
    if not best or best[0] < 3:
        raise ValueError('PRODUCT_WORKBOOK_HEADER_NOT_FOUND')
    return best[1], best[2]


def _matches_identity(value: Any, identity: CanonicalIdentity) -> bool:
    text = _text(value).casefold()
    if not text:
        return False
    candidates = [
        identity.manufacturer_part_number,
        identity.commercial_model,
        *identity.ean_upc_gtin,
    ]
    return any(text == _text(candidate).casefold() for candidate in candidates if _text(candidate))


def _choose_target_row(ws, header_row: int, columns: dict[str, int], identity: CanonicalIdentity) -> int:
    populated: list[int] = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        row_has_value = any(_text(ws.cell(row_idx, col_idx).value) for col_idx in range(1, ws.max_column + 1))
        if not row_has_value:
            continue
        populated.append(row_idx)
        for key in ('sku', 'model', 'barcode'):
            col = columns.get(key)
            if col and _matches_identity(ws.cell(row_idx, col).value, identity):
                return row_idx
    if populated:
        return populated[0]
    return header_row + 1


def _purge_conflicting_rows(ws, header_row: int, target_row: int, columns: dict[str, int], identity: CanonicalIdentity) -> None:
    for row_idx in range(header_row + 1, ws.max_row + 1):
        if row_idx == target_row:
            continue
        identity_values = [
            _text(ws.cell(row_idx, columns[key]).value)
            for key in ('brand', 'sku', 'model', 'barcode')
            if columns.get(key)
        ]
        if not any(identity_values):
            continue
        if any(_matches_identity(value, identity) for value in identity_values):
            continue
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row_idx, col_idx).value = None


def _first_identity_source(identity: CanonicalIdentity) -> tuple[str, str]:
    for source in identity.sources:
        if not isinstance(source, dict):
            continue
        url = _text(source.get('url') or source.get('source_url') or source.get('fuente'))
        source_type = _text(source.get('source_type') or source.get('tipo_fuente') or 'OFFICIAL_PRODUCT')
        if url or source_type:
            return url, source_type
    return '', 'IDENTITY'


def _clean_sentinel(value: Any) -> Any:
    return '' if _text(value) == '89' else value


def _old_evidence_rows(wb) -> list[list[Any]]:
    if 'IA_EVIDENCIA' not in wb.sheetnames:
        return []
    ws = wb['IA_EVIDENCIA']
    if ws.max_row < 2:
        return []
    headers = [_norm(cell.value) for cell in ws[1]]
    positions = {header: index for index, header in enumerate(headers)}
    result: list[list[Any]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        def get(*aliases: str):
            for alias in aliases:
                idx = positions.get(_norm(alias))
                if idx is not None and idx < len(values):
                    return values[idx]
            return ''
        part_number = _clean_sentinel(get('Part Number'))
        field_name = _clean_sentinel(get('Campo', 'Campo plantilla'))
        proposed = _clean_sentinel(get('Valor', 'Valor propuesto'))
        status_raw = _text(get('Estado')).upper()
        status = {'CONFIRMADO': 'CONFIRMED', 'RECHAZADO_LOCAL': 'REJECTED'}.get(status_raw, status_raw)
        written = proposed if status == 'CONFIRMED' and _text(proposed) else ''
        row = [
            part_number,
            field_name,
            proposed,
            written,
            status,
            _clean_sentinel(get('Confianza')),
            _clean_sentinel(get('Fuente')),
            _clean_sentinel(get('Tipo fuente')),
            _clean_sentinel(get('Página PDF', 'Pagina PDF')),
            _clean_sentinel(get('Evidencia')),
            _clean_sentinel(get('Observación', 'Observacion')),
        ]
        if any(_text(value) for value in row):
            result.append(row)
    return result


def _write_master_sheet(
    wb,
    identity: CanonicalIdentity,
    specifications: list[MasterSpecification],
    old_rows: list[list[Any]],
) -> None:
    if 'ESPECIFICACIONES_COMPLETAS' in wb.sheetnames:
        del wb['ESPECIFICACIONES_COMPLETAS']
    ws = wb.create_sheet('ESPECIFICACIONES_COMPLETAS')
    ws.append(_MASTER_HEADERS)
    seen: set[tuple[str, str]] = set()
    for spec in specifications:
        label = spec.label or spec.key
        ws.append([
            identity.manufacturer_part_number,
            '',
            label,
            spec.value,
            spec.unit,
            spec.status,
            spec.confidence,
            spec.source_url,
            spec.source_type,
            spec.pdf_page,
            spec.evidence,
        ])
        seen.add((_norm(label), _norm(spec.value)))
    for row in old_rows:
        if len(row) < 11 or _text(row[4]).upper() != 'CONFIRMED':
            continue
        label = _text(row[1])
        value = _text(row[3] or row[2])
        if not label or not value or value == '89':
            continue
        key = (_norm(label), _norm(value))
        if key in seen:
            continue
        seen.add(key)
        ws.append([
            identity.manufacturer_part_number or row[0],
            '',
            label,
            value,
            '',
            'CONFIRMED',
            row[5],
            row[6],
            row[7],
            row[8],
            row[9] or row[10],
        ])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = 'A2'


def _write_evidence_sheet(
    wb,
    identity: CanonicalIdentity,
    specifications: list[MasterSpecification],
    evidence_rows: list[dict],
    old_rows: list[list[Any]],
) -> None:
    if 'IA_EVIDENCIA' in wb.sheetnames:
        del wb['IA_EVIDENCIA']
    ws = wb.create_sheet('IA_EVIDENCIA')
    ws.append(_EVIDENCE_HEADERS)
    for row in old_rows:
        ws.append(row)

    source_url, source_type = _first_identity_source(identity)
    identity_entries = [
        ('Marca #26', identity.brand),
        ('SKU del vendedor #29', identity.manufacturer_part_number),
        ('Modelo #32', identity.commercial_model),
    ]
    if identity.ean_upc_gtin:
        identity_entries.append(('Código de barras #56', identity.ean_upc_gtin[0]))
    existing_pairs = {(_text(row[0]).casefold(), _text(row[1]).casefold(), _text(row[3]).casefold()) for row in old_rows}
    for field_name, value in identity_entries:
        if not _text(value):
            continue
        pair = (identity.manufacturer_part_number.casefold(), field_name.casefold(), _text(value).casefold())
        if pair in existing_pairs:
            continue
        ws.append([
            identity.manufacturer_part_number,
            field_name,
            value,
            value,
            'CONFIRMED',
            identity.confidence,
            source_url,
            source_type,
            None,
            'Identidad canónica validada por Product Data Intelligence.',
            '',
        ])

    for spec in specifications:
        ws.append([
            identity.manufacturer_part_number,
            spec.label or spec.key,
            spec.value,
            spec.value if spec.status == 'CONFIRMED' else '',
            spec.status,
            spec.confidence,
            spec.source_url,
            spec.source_type,
            spec.pdf_page,
            spec.evidence,
            '',
        ])

    for item in evidence_rows:
        if not isinstance(item, dict):
            continue
        ws.append([
            identity.manufacturer_part_number,
            _clean_sentinel(item.get('field') or item.get('campo')),
            _clean_sentinel(item.get('proposed') or item.get('valor_propuesto') or item.get('value')),
            _clean_sentinel(item.get('written') or item.get('valor_escrito')),
            _clean_sentinel(item.get('status') or item.get('estado')),
            _clean_sentinel(item.get('confidence') or item.get('confianza')),
            _clean_sentinel(item.get('source_url') or item.get('fuente')),
            _clean_sentinel(item.get('source_type') or item.get('tipo_fuente')),
            _clean_sentinel(item.get('pdf_page') or item.get('pagina_pdf')),
            _clean_sentinel(item.get('evidence') or item.get('evidencia')),
            _clean_sentinel(item.get('observation') or item.get('observacion')),
        ])

    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = 'A2'


def finalize_product_workbook(
    path: Path,
    identity: CanonicalIdentity,
    specifications: list[MasterSpecification],
    evidence_rows: list[dict],
) -> ProductWorkbookQA:
    path = Path(path)
    wb = load_workbook(path)
    errors: list[str] = []
    warnings: list[str] = []
    target_row: int | None = None
    try:
        if 'Subir plantilla' not in wb.sheetnames:
            errors.append('MISSING_SHEET:Subir plantilla')
        else:
            ws = wb['Subir plantilla']
            header_row, columns = _find_header_row(ws)
            target_row = _choose_target_row(ws, header_row, columns, identity)
            _purge_conflicting_rows(ws, header_row, target_row, columns, identity)

            critical = {
                'brand': identity.brand,
                'sku': identity.manufacturer_part_number,
                'model': identity.commercial_model,
            }
            for key, value in critical.items():
                if not _text(value):
                    errors.append(f'MISSING_IDENTITY:{key}')
                    continue
                col = columns.get(key)
                if not col:
                    errors.append(f'MISSING_COLUMN:{key}')
                    continue
                ws.cell(target_row, col).value = value

            barcode_col = columns.get('barcode')
            if barcode_col and identity.ean_upc_gtin:
                ws.cell(target_row, barcode_col).value = identity.ean_upc_gtin[0]

            old_rows = _old_evidence_rows(wb)
            _write_master_sheet(wb, identity, specifications, old_rows)
            _write_evidence_sheet(wb, identity, specifications, evidence_rows, old_rows)

            for key, expected in critical.items():
                col = columns.get(key)
                if col and _text(expected):
                    actual = _text(ws.cell(target_row, col).value)
                    if actual.casefold() != _text(expected).casefold():
                        errors.append(f'IDENTITY_MISMATCH:{key}:{actual}')

            for row_idx in range(header_row + 1, ws.max_row + 1):
                if row_idx == target_row:
                    continue
                for key in ('sku', 'model', 'barcode'):
                    col = columns.get(key)
                    value = _text(ws.cell(row_idx, col).value) if col else ''
                    if value and not _matches_identity(value, identity):
                        errors.append(f'CROSS_PRODUCT_ROW:{row_idx}:{key}:{value}')

        wb.save(path)
    finally:
        wb.close()
    return ProductWorkbookQA(ok=not errors, errors=errors, warnings=warnings, target_row=target_row)


def assert_product_workbook_qa(qa: ProductWorkbookQA) -> None:
    if not qa.ok:
        raise ValueError('PRODUCT_WORKBOOK_QA_FAILED: ' + '; '.join(qa.errors))

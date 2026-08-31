from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any
import shutil
import re
import unicodedata

from openpyxl import load_workbook
from openpyxl.styles import Font

from .marketplace_prompt_contract import validate_template_value
from .marketplace_template import MarketplaceTemplateProfile, ProductSlot, TemplateField, is_placeholder_value
from .product_identity import CanonicalIdentity
from .product_evidence import MasterSpecification

_ACCEPTED_STATUSES = {'CONFIRMADO', 'CONFIRMED', 'ACCEPTED', 'ACEPTADO', 'VALIDATED', 'VALIDADO'}
_EVIDENCE_HEADERS = [
    'Part Number', 'Campo plantilla', 'Valor propuesto', 'Valor escrito', 'Estado',
    'Confianza', 'Fuente', 'Tipo fuente', 'Página PDF', 'Evidencia', 'Observación',
    'Marketplace', 'Hoja origen', 'Fila origen',
]
_MASTER_HEADERS = [
    'Part Number', 'Categoría técnica', 'Especificación', 'Valor', 'Unidad', 'Estado',
    'Confianza', 'Fuente principal', 'Tipo fuente', 'Página PDF', 'Observación',
    'Origen', 'Marketplace', 'Hoja origen', 'Fila origen',
]
_PRODUCT_HEADERS = [
    'Marketplace', 'Hoja', 'Fila', 'Identificador entrada', 'Tipo identificador',
    'Part Number', 'Modelo comercial', 'Marca', 'EAN/UPC/GTIN', 'Estado QA',
    'Campos escritos', 'Campos preservados', 'Requeridos faltantes', 'Advertencias',
]


def _text(value: Any) -> str:
    return str(value or '').strip()


def _norm(value: Any) -> str:
    text = unicodedata.normalize('NFKD', _text(value).casefold())
    text = ''.join(ch for ch in text if not unicodedata.combining(ch))
    return ' '.join(text.replace('_', ' ').split())


def _compact(value: Any) -> str:
    return re.sub(r'[^a-z0-9]+', '', _norm(value))


def _item_value(item: Any, name: str, default: Any = '') -> Any:
    if isinstance(item, dict):
        return item.get(name, default)
    return getattr(item, name, default)


def _preview_status_ok(item: Any) -> bool:
    return _text(_item_value(item, 'status')).upper() in _ACCEPTED_STATUSES


def _match_field(profile: MarketplaceTemplateProfile, name: str) -> TemplateField | None:
    direct = profile.field_by_label(name)
    if direct:
        return direct
    direct = profile.field_by_code(name)
    if direct:
        return direct
    target = _compact(re.sub(r'\s*#\s*\d+\s*$', '', _text(name)))
    for field in profile.fields:
        if _compact(re.sub(r'\s*#\s*\d+\s*$', '', field.label)) == target:
            return field
        if _compact(field.code) == target:
            return field
    return None


def _identity_source(identity: CanonicalIdentity) -> tuple[str, str]:
    for source in identity.sources or []:
        if not isinstance(source, dict):
            continue
        url = _text(source.get('url') or source.get('source_url') or source.get('fuente'))
        kind = _text(source.get('source_type') or source.get('tipo_fuente') or 'IDENTITY')
        if url or kind:
            return url, kind
    return '', 'IDENTITY'


@dataclass
class ProductWriteRecord:
    slot: ProductSlot
    identity: CanonicalIdentity
    preview_rows: list[Any] = field(default_factory=list)
    specifications: list[MasterSpecification] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


@dataclass
class ProductWriteSummary:
    row: int
    identifier: str
    part_number: str
    ok: bool
    written: int
    preserved: int
    missing_required: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


@dataclass
class MarketplaceWorkbookQA:
    ok: bool
    products: list[ProductWriteSummary] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


def _set_if_blank_or_example(ws, row: int, field: TemplateField | None, value: Any, *, force: bool = False) -> tuple[bool, bool]:
    if field is None or not _text(value):
        return False, False
    cell = ws.cell(row, field.column)
    current = cell.value
    replaceable = not _text(current) or is_placeholder_value(field, current)
    if force or replaceable:
        if _text(current) != _text(value):
            cell.value = value
            return True, False
        return False, False
    return False, True


def _field_value(ws, row: int, field: TemplateField | None) -> str:
    if field is None:
        return ''
    return _text(ws.cell(row, field.column).value)


def _semantic_fields(profile: MarketplaceTemplateProfile) -> dict[str, TemplateField | None]:
    if profile.marketplace == 'falabella':
        return {
            'brand': profile.field_by_label('Marca'),
            'mpn': profile.field_by_label('SKU del vendedor'),
            'model': profile.field_by_label('Modelo'),
            'barcode': profile.field_by_label('Código de barras', 'Codigo de barras'),
            'category': profile.field_by_label('Categoría primaria'),
            'condition': profile.field_by_label('Condición del Producto'),
            'description': profile.field_by_label('Descripción', 'Descripcion'),
        }
    return {
        'brand': profile.field_by_code('marca'),
        'mpn': None,
        'model': profile.field_by_code('modelo'),
        'barcode': profile.field_by_code('ean'),
        'category': profile.field_by_code('categoria'),
        'condition': profile.field_by_code('state'),
        'description': profile.field_by_code('descripcion'),
    }


def _safe_condition_value(field: TemplateField | None) -> str:
    if field is None:
        return ''
    for value in field.valid_values:
        if _norm(value) == 'nuevo':
            return value
    return ''


def _append_header(ws, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = 'A2'


def _collect_existing_sheet_rows(wb, name: str, width: int) -> list[list[Any]]:
    if name not in wb.sheetnames:
        return []
    ws = wb[name]
    result: list[list[Any]] = []
    if (ws.max_row or 0) < 2:
        return result
    for row in ws.iter_rows(min_row=2, values_only=True):
        values = list(row[:width])
        while len(values) < width:
            values.append('')
        if any(_text(value) for value in values):
            result.append(values)
    return result


def _prepare_trace_sheets(wb):
    old_evidence = _collect_existing_sheet_rows(wb, 'IA_EVIDENCIA', len(_EVIDENCE_HEADERS))
    old_master = _collect_existing_sheet_rows(wb, 'ESPECIFICACIONES_COMPLETAS', len(_MASTER_HEADERS))
    for name in ('IA_PRODUCTOS', 'IA_EVIDENCIA', 'ESPECIFICACIONES_COMPLETAS'):
        if name in wb.sheetnames:
            del wb[name]
    products = wb.create_sheet('IA_PRODUCTOS')
    evidence = wb.create_sheet('IA_EVIDENCIA')
    master = wb.create_sheet('ESPECIFICACIONES_COMPLETAS')
    _append_header(products, _PRODUCT_HEADERS)
    _append_header(evidence, _EVIDENCE_HEADERS)
    _append_header(master, _MASTER_HEADERS)
    for row in old_evidence:
        evidence.append(row)
    for row in old_master:
        master.append(row)
    return products, evidence, master


def _append_identity_evidence(ws, profile: MarketplaceTemplateProfile, record: ProductWriteRecord):
    identity = record.identity
    source_url, source_type = _identity_source(identity)
    entries = [
        ('Marca', identity.brand),
        ('Part Number / MPN', identity.manufacturer_part_number),
        ('Modelo comercial', identity.commercial_model),
    ]
    if identity.ean_upc_gtin:
        entries.append(('EAN/UPC/GTIN', identity.ean_upc_gtin[0]))
    for label, value in entries:
        if not _text(value):
            continue
        ws.append([
            identity.manufacturer_part_number,
            label,
            value,
            value,
            'CONFIRMED',
            identity.confidence,
            source_url,
            source_type,
            None,
            'Identidad canónica del producto.',
            '',
            profile.marketplace,
            profile.sheet_name,
            record.slot.row,
        ])


def _append_preview_trace(evidence_ws, master_ws, profile: MarketplaceTemplateProfile, record: ProductWriteRecord):
    mpn = record.identity.manufacturer_part_number or record.slot.identifier
    for item in record.preview_rows:
        field_name = _text(_item_value(item, 'field'))
        value = _text(_item_value(item, 'value'))
        status = _text(_item_value(item, 'status')).upper()
        confidence = _item_value(item, 'confidence', 0)
        reason = _text(_item_value(item, 'reason'))
        if not field_name and not value:
            continue
        written = value if status in _ACCEPTED_STATUSES else ''
        evidence_ws.append([
            mpn, field_name, value, written, status, confidence, '', 'MAPPED_FIELD', None,
            reason, '', profile.marketplace, profile.sheet_name, record.slot.row,
        ])
        if status in _ACCEPTED_STATUSES and value:
            master_ws.append([
                mpn, 'MAPEADO_PLANTILLA', field_name, value, '', 'CONFIRMED', confidence,
                '', 'MAPPED_FIELD', None, reason, 'MAPEADO', profile.marketplace,
                profile.sheet_name, record.slot.row,
            ])

    for spec in record.specifications:
        status = _text(spec.status).upper()
        if not _text(spec.value):
            continue
        master_ws.append([
            mpn,
            'EXTRA',
            spec.label or spec.key,
            spec.value,
            spec.unit,
            status,
            spec.confidence,
            spec.source_url,
            spec.source_type,
            spec.pdf_page,
            spec.evidence,
            'INVESTIGADO_ADICIONAL',
            profile.marketplace,
            profile.sheet_name,
            record.slot.row,
        ])
        evidence_ws.append([
            mpn,
            spec.label or spec.key,
            spec.value,
            spec.value if status == 'CONFIRMED' else '',
            status,
            spec.confidence,
            spec.source_url,
            spec.source_type,
            spec.pdf_page,
            spec.evidence,
            'Especificación adicional investigada.',
            profile.marketplace,
            profile.sheet_name,
            record.slot.row,
        ])


def _responsible_field_keys(profile: MarketplaceTemplateProfile, record: ProductWriteRecord, semantic: dict[str, TemplateField | None]) -> set[str]:
    keys: set[str] = set()
    for field in semantic.values():
        if field is not None:
            keys.add(_compact(field.code))
    for item in record.preview_rows:
        matched = _match_field(profile, _item_value(item, 'field'))
        if matched is not None:
            keys.add(_compact(matched.code))
    return keys


def _missing_required(
    ws,
    profile: MarketplaceTemplateProfile,
    slot: ProductSlot,
    responsible_keys: set[str],
    validated_keys: set[str],
) -> list[str]:
    missing: list[str] = []
    for field in profile.fields:
        if _compact(field.code) not in responsible_keys:
            continue
        if field.requirement_for(slot.category).upper() not in {'REQUIRED', 'MANDATORY'}:
            continue
        value = ws.cell(slot.row, field.column).value
        key = _compact(field.code)
        if not _text(value) or (is_placeholder_value(field, value) and key not in validated_keys):
            missing.append(field.label or field.code)
    return missing


def write_marketplace_workbook(
    source_path: Path,
    output_path: Path,
    profile: MarketplaceTemplateProfile,
    records: list[ProductWriteRecord],
) -> MarketplaceWorkbookQA:
    source_path = Path(source_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if source_path.resolve() != output_path.resolve():
        shutil.copy2(source_path, output_path)

    wb = load_workbook(output_path)
    errors: list[str] = []
    warnings: list[str] = []
    summaries: list[ProductWriteSummary] = []
    try:
        if profile.sheet_name not in wb.sheetnames:
            return MarketplaceWorkbookQA(False, errors=[f'MISSING_SHEET:{profile.sheet_name}'])
        ws = wb[profile.sheet_name]
        products_ws, evidence_ws, master_ws = _prepare_trace_sheets(wb)
        semantic = _semantic_fields(profile)

        record_by_row = {record.slot.row: record for record in records}
        expected_rows = {slot.row for slot in profile.products}
        missing_records = sorted(expected_rows - set(record_by_row))
        for row in missing_records:
            errors.append(f'UNPROCESSED_PRODUCT_ROW:{row}')

        for record in records:
            slot = record.slot
            identity = record.identity
            written = 0
            preserved = 0
            local_warnings = list(record.warnings)
            validated_keys: set[str] = set()

            changed, kept = _set_if_blank_or_example(ws, slot.row, semantic['brand'], identity.brand, force=bool(identity.brand))
            written += int(changed); preserved += int(kept)
            if identity.brand and semantic['brand'] is not None:
                validated_keys.add(_compact(semantic['brand'].code))
            if profile.marketplace == 'falabella':
                changed, kept = _set_if_blank_or_example(ws, slot.row, semantic['mpn'], identity.manufacturer_part_number, force=bool(identity.manufacturer_part_number))
                written += int(changed); preserved += int(kept)
                if identity.manufacturer_part_number and semantic['mpn'] is not None:
                    validated_keys.add(_compact(semantic['mpn'].code))
            model_value = identity.commercial_model or identity.manufacturer_part_number
            changed, kept = _set_if_blank_or_example(ws, slot.row, semantic['model'], model_value, force=bool(model_value))
            written += int(changed); preserved += int(kept)
            if model_value and semantic['model'] is not None:
                validated_keys.add(_compact(semantic['model'].code))
            if identity.ean_upc_gtin:
                changed, kept = _set_if_blank_or_example(ws, slot.row, semantic['barcode'], identity.ean_upc_gtin[0], force=True)
                written += int(changed); preserved += int(kept)
                if semantic['barcode'] is not None:
                    validated_keys.add(_compact(semantic['barcode'].code))

            if semantic['category'] is not None and len(profile.category_options) == 1:
                changed, kept = _set_if_blank_or_example(ws, slot.row, semantic['category'], profile.category_options[0])
                written += int(changed); preserved += int(kept)
                validated_keys.add(_compact(semantic['category'].code))
            condition_value = _safe_condition_value(semantic['condition'])
            if condition_value:
                changed, kept = _set_if_blank_or_example(ws, slot.row, semantic['condition'], condition_value)
                written += int(changed); preserved += int(kept)
                validated_keys.add(_compact(semantic['condition'].code))

            for item in record.preview_rows:
                if not _preview_status_ok(item):
                    continue
                value = _item_value(item, 'value')
                if not _text(value):
                    continue
                field = _match_field(profile, _item_value(item, 'field'))
                if field is None:
                    continue
                identity_fields = (semantic['brand'], semantic['mpn'], semantic['model'], semantic['barcode'])
                if any(field is identity_field for identity_field in identity_fields):
                    continue
                check = validate_template_value(field, value)
                if not check.ok:
                    local_warnings.append(
                        f'TEMPLATE_VALUE_REJECTED:{field.label or field.code}:{check.reason}'
                    )
                    continue
                changed, kept = _set_if_blank_or_example(ws, slot.row, field, check.value)
                written += int(changed); preserved += int(kept)
                validated_keys.add(_compact(field.code))

            description = _field_value(ws, slot.row, semantic['description'])
            if description and len(description) < 350:
                local_warnings.append(f'DESCRIPTION_TOO_SHORT:{len(description)}')

            responsible_keys = _responsible_field_keys(profile, record, semantic)
            missing = _missing_required(ws, profile, slot, responsible_keys, validated_keys)
            product_ok = bool(identity.manufacturer_part_number or slot.identifier) and not missing
            if not identity.brand:
                local_warnings.append('MISSING_IDENTITY:brand')
            if not identity.commercial_model:
                local_warnings.append('MISSING_IDENTITY:commercial_model')
            if missing:
                errors.append(f'ROW_{slot.row}_MISSING_REQUIRED:' + '|'.join(missing))

            _append_identity_evidence(evidence_ws, profile, record)
            _append_preview_trace(evidence_ws, master_ws, profile, record)

            summary = ProductWriteSummary(
                row=slot.row,
                identifier=slot.identifier,
                part_number=identity.manufacturer_part_number,
                ok=product_ok,
                written=written,
                preserved=preserved,
                missing_required=missing,
                warnings=local_warnings,
            )
            summaries.append(summary)
            products_ws.append([
                profile.marketplace,
                profile.sheet_name,
                slot.row,
                slot.identifier,
                slot.identifier_type,
                identity.manufacturer_part_number,
                identity.commercial_model,
                identity.brand,
                identity.ean_upc_gtin[0] if identity.ean_upc_gtin else '',
                'COMPLETADO' if product_ok else 'NO_VALIDADO',
                written,
                preserved,
                ' | '.join(missing),
                ' | '.join(local_warnings),
            ])

        wb.save(output_path)
    finally:
        wb.close()

    return MarketplaceWorkbookQA(ok=not errors, products=summaries, errors=errors, warnings=warnings)

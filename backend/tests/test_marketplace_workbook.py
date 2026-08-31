import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.marketplace_template import analyze_marketplace_template
from app.marketplace_workbook import ProductWriteRecord, write_marketplace_workbook
from app.product_identity import CanonicalIdentity
from app.product_evidence import MasterSpecification


def make_falabella(path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Subir plantilla'
    ws.append(['Principales'] * 9)
    ws.append([
        '- Value: Esto es un párrafo', '- Value: EPSON', '- Value: ABC-100', '- Value: ABC-100',
        '- Value: 1234567890123', '- Value: Esto es un párrafo', '- Value: 888 - Impresoras',
        '- Value: Nuevo', '- Value: 999.99',
    ])
    ws.append([' ', ' ', ' ', ' ', '( Optional )', ' ', ' ', ' ', ' '])
    ws.append([
        'Nombre #39', 'Marca #26', 'SKU del vendedor #29', 'Modelo #32', 'Código de barras #56',
        'Descripción #53', 'Categoría primaria #1', 'Condición del Producto #22', 'Precio #10',
    ])
    ws.append(['', '', '', 'C11CL62301', '', '', '', '', 777.00])
    ws.append(['Nombre ya válido', '', '', 'C11CL65301', '', '', '', '', 888.00])
    cats = wb.create_sheet('Categorías')
    cats.append(['Categorías']); cats.append(['888 - Impresoras'])
    opts = wb.create_sheet('Opciones')
    opts.append(['ConditionType']); opts.append(['Nuevo']); opts.append(['Open Box'])
    wb.save(path)


def make_ripley(path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Data'
    ws.append(['Categoría','Nombre','EAN/UPC','Marca','Modelo','Descripción','SKU Seller','Estado','Precio','Cantidad','Product ID'])
    ws.append(['categoria','nombre','ean','marca','modelo','descripcion','sku_seller','state','price','quantity','product_id'])
    ws.append(['CAT-MOUSE','Mouse existente','1234567890123','','RGBM-02-BK','','SELLER-A','Nuevo',49.90,7,'RID-001'])
    cols = wb.create_sheet('Columns')
    cols.append(['Código','Etiqueta','Descripción','Valor de ejemplo','CAT-MOUSE'])
    for code,label,example,status in [
        ('categoria','Categoría','','REQUIRED'),('nombre','Nombre','Nombre demo','REQUIRED'),
        ('ean','EAN/UPC','1234567890123','RECOMMENDED'),('marca','Marca','Marca demo','REQUIRED'),
        ('modelo','Modelo','ABC-100','REQUIRED'),('descripcion','Descripción','Texto demo','REQUIRED'),
        ('sku_seller','SKU Seller','SELLER-DEMO','REQUIRED'),('state','Estado','','REQUIRED'),
        ('price','Precio','','REQUIRED'),('quantity','Cantidad','','REQUIRED'),('product_id','Product ID','','REQUIRED'),
    ]:
        cols.append([code,label,'',example,status])
    ref = wb.create_sheet('ReferenceData')
    ref.append(['state']); ref.append(['Nuevo']); ref.append(['Usado'])
    wb.save(path)


def identity(mpn, model, brand='EPSON', barcode=''):
    return CanonicalIdentity(
        brand=brand,
        manufacturer_part_number=mpn,
        commercial_model=model,
        ean_upc_gtin=[barcode] if barcode else [],
        confidence=99,
        sources=[{'url': 'https://official.example/product', 'source_type': 'OFFICIAL_PRODUCT'}],
    )


class MarketplaceWorkbookTests(unittest.TestCase):
    def test_two_falabella_products_are_updated_in_place_without_purging_other_row(self):
        with tempfile.TemporaryDirectory() as td:
            src = Path(td) / 'source.xlsx'
            out = Path(td) / 'result.xlsx'
            make_falabella(src)
            profile = analyze_marketplace_template(src)
            records = [
                ProductWriteRecord(
                    slot=profile.products[0],
                    identity=identity('C11CL62301','L3350',barcode='198390011567'),
                    preview_rows=[
                        {'field':'Nombre #39','value':'Epson EcoTank L3350 Multifuncional Wi-Fi','status':'CONFIRMED','confidence':99},
                        {'field':'Descripción #53','value':'D' * 700,'status':'CONFIRMED','confidence':98},
                    ],
                    specifications=[MasterSpecification('resolution','Resolución','5760 x 1440','dpi','CONFIRMED',97,'https://official.example/spec.pdf','OFFICIAL_PDF','Ficha',2,'Resolución oficial','C11CL62301')],
                ),
                ProductWriteRecord(
                    slot=profile.products[1],
                    identity=identity('C11CL65301','L1350',barcode='0103439999999'),
                    preview_rows=[
                        {'field':'Nombre #39','value':'Nombre propuesto que no debe pisar el existente','status':'CONFIRMED','confidence':99},
                        {'field':'Descripción #53','value':'E' * 700,'status':'CONFIRMED','confidence':98},
                    ],
                ),
            ]
            qa = write_marketplace_workbook(src, out, profile, records)

            self.assertTrue(qa.ok, qa.errors)
            wb = load_workbook(out)
            ws = wb['Subir plantilla']
            self.assertEqual(ws['C5'].value, 'C11CL62301')
            self.assertEqual(ws['D5'].value, 'L3350')
            self.assertEqual(ws['C6'].value, 'C11CL65301')
            self.assertEqual(ws['D6'].value, 'L1350')
            self.assertEqual(ws['A6'].value, 'Nombre ya válido')
            self.assertEqual(ws['I5'].value, 777.00)
            self.assertEqual(ws['I6'].value, 888.00)
            self.assertEqual(ws['G5'].value, '888 - Impresoras')
            self.assertEqual(ws['H5'].value, 'Nuevo')
            self.assertEqual(ws['G6'].value, '888 - Impresoras')
            self.assertEqual(ws['H6'].value, 'Nuevo')
            self.assertEqual(wb['IA_PRODUCTOS'].max_row, 3)
            master_values = [row[2].value for row in wb['ESPECIFICACIONES_COMPLETAS'].iter_rows(min_row=2)]
            self.assertIn('Resolución', master_values)
            evidence_mpns = {row[0].value for row in wb['IA_EVIDENCIA'].iter_rows(min_row=2)}
            self.assertIn('C11CL62301', evidence_mpns)
            self.assertIn('C11CL65301', evidence_mpns)
            wb.close()

    def test_ripley_preserves_operational_offer_fields_and_fills_only_characteristics(self):
        with tempfile.TemporaryDirectory() as td:
            src = Path(td) / 'ripley.xlsx'
            out = Path(td) / 'result.xlsx'
            make_ripley(src)
            profile = analyze_marketplace_template(src)
            record = ProductWriteRecord(
                slot=profile.products[0],
                identity=identity('RGBM-02-BK','RGBM-02-BK',brand='TEROS',barcode='1234567890123'),
                preview_rows=[
                    {'field':'Nombre','value':'Mouse TEROS RGBM-02-BK','status':'CONFIRMED','confidence':95},
                    {'field':'Descripción','value':'Mouse verificado. ' + ('X' * 680),'status':'CONFIRMED','confidence':95},
                ],
            )
            qa = write_marketplace_workbook(src, out, profile, [record])
            self.assertTrue(qa.ok, qa.errors)
            wb = load_workbook(out)
            ws = wb['Data']
            self.assertEqual(ws['I3'].value, 49.90)
            self.assertEqual(ws['J3'].value, 7)
            self.assertEqual(ws['K3'].value, 'RID-001')
            self.assertEqual(ws['G3'].value, 'SELLER-A')
            self.assertEqual(ws['D3'].value, 'TEROS')
            self.assertTrue(str(ws['F3'].value).startswith('Mouse verificado.'))
            wb.close()

    def test_unprocessed_detected_row_blocks_workbook_qa_but_does_not_delete_it(self):
        with tempfile.TemporaryDirectory() as td:
            src = Path(td) / 'source.xlsx'
            out = Path(td) / 'result.xlsx'
            make_falabella(src)
            profile = analyze_marketplace_template(src)
            record = ProductWriteRecord(
                slot=profile.products[0],
                identity=identity('C11CL62301','L3350'),
                preview_rows=[
                    {'field':'Nombre #39','value':'Producto 1','status':'CONFIRMED','confidence':99},
                    {'field':'Descripción #53','value':'D' * 700,'status':'CONFIRMED','confidence':99},
                ],
            )
            qa = write_marketplace_workbook(src, out, profile, [record])
            self.assertFalse(qa.ok)
            self.assertIn('UNPROCESSED_PRODUCT_ROW:6', qa.errors)
            wb = load_workbook(out)
            self.assertEqual(wb['Subir plantilla']['D6'].value, 'C11CL65301')
            wb.close()


if __name__ == '__main__':
    unittest.main()

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from app.product_characteristics import build_product_intelligence, resolve_characteristics_input


def workbook_with_stale_model_and_evidence(path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Subir plantilla'
    ws.append(['Marca #26', 'SKU del vendedor #29', 'Modelo #32', 'Código de barras #56'])
    ws.append(['', '', 'jblq350wlblkam', ''])
    ev = wb.create_sheet('IA_EVIDENCIA')
    ev.append(['Part Number', 'Campo', 'Valor', 'Estado', 'Confianza', 'Fuente', 'Tipo fuente'])
    ev.append(['C11CL62301', 'Marca #26', 'EPSON', 'CONFIRMADO', 100, 'https://epson.test/product', 'OFICIAL'])
    ev.append(['C11CL62301', 'Modelo #32', 'L3350', 'CONFIRMADO', 99, 'https://epson.test/support', 'SOPORTE'])
    wb.save(path)


class ProductCharacteristicsTests(unittest.TestCase):
    def test_resolves_manual_or_auto_input_from_same_workbook(self):
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / 'template.xlsx'
            workbook_with_stale_model_and_evidence(path)
            manual = resolve_characteristics_input(path, 'C11CL62301')
            auto = resolve_characteristics_input(path, '')
        self.assertEqual((manual.input_mode, manual.identifier), ('manual', 'C11CL62301'))
        self.assertEqual((auto.input_mode, auto.identifier), ('auto', 'C11CL62301'))
        self.assertEqual(auto.identifier_type, 'PART_NUMBER')

    def test_builds_canonical_identity_and_accepts_official_master_spec(self):
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / 'template.xlsx'
            workbook_with_stale_model_and_evidence(path)
            result = build_product_intelligence({
                'producto': {'marca': 'EPSON', 'modelo': 'L3350', 'confidence': 98},
                'especificaciones_completas': [{
                    'key': 'resolution', 'label': 'Resolución máxima', 'value': '5760 x 1440',
                    'unit': 'dpi', 'status': 'CONFIRMED', 'confidence': 97,
                    'source_url': 'https://epson.test/spec.pdf', 'source_type': 'OFFICIAL_PDF',
                    'source_title': 'Ficha técnica', 'pdf_page': 2,
                    'evidence': '5760 x 1440 dpi', 'applies_to': 'C11CL62301',
                }],
            }, path, 'C11CL62301')
        self.assertTrue(result.qa_ready, result.critical_errors)
        self.assertEqual(result.identity.manufacturer_part_number, 'C11CL62301')
        self.assertEqual(result.identity.brand, 'EPSON')
        self.assertEqual(result.identity.commercial_model, 'L3350')
        self.assertEqual(len(result.specifications), 1)
        self.assertEqual(result.specifications[0].source_type, 'OFFICIAL_PDF')

    def test_missing_commercial_model_blocks_qa_ready(self):
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / 'template.xlsx'
            wb = Workbook()
            ws = wb.active
            ws.title = 'Subir plantilla'
            ws.append(['Marca #26', 'SKU del vendedor #29', 'Modelo #32'])
            ws.append(['EPSON', '', 'C11CL62301'])
            wb.save(path)
            result = build_product_intelligence({'producto': {'marca': 'EPSON'}}, path, 'C11CL62301')
        self.assertFalse(result.qa_ready)
        self.assertIn('MISSING_IDENTITY:commercial_model', result.critical_errors)


if __name__ == '__main__':
    unittest.main()

import sys
import tempfile
import types
import unittest
from contextlib import contextmanager
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

from openpyxl import Workbook, load_workbook


def make_workbook(path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Subir plantilla'
    ws.append(['Marca #26','SKU del vendedor #29','Modelo #32','Código de barras #56'])
    ws.append(['','','jblq350wlblkam',''])
    ev = wb.create_sheet('IA_EVIDENCIA')
    ev.append(['Part Number','Campo','Valor','Estado','Confianza','Fuente','Tipo fuente'])
    ev.append(['C11CL62301','Marca #26','EPSON','CONFIRMADO',100,'https://epson.test/product','OFICIAL'])
    ev.append(['C11CL62301','Modelo #32','L3350','CONFIRMADO',99,'https://epson.test/support','SOPORTE'])
    wb.save(path)


class FakeSession:
    async def __aenter__(self): return self
    async def __aexit__(self, *args): return False
    async def ask(self, *args, **kwargs): return '{}'


def _module(name: str, **attrs):
    module = types.ModuleType(name)
    for key, value in attrs.items():
        setattr(module, key, value)
    return module


@contextmanager
def isolated_workflows(
    *,
    prepare_research=None,
    run_research_for_preview_once=None,
    write_validated_workbook=None,
):
    """Import app.workflows against temporary legacy stubs and restore sys.modules exactly."""
    default_schema = SimpleNamespace(
        family='falabella', sheet_name='Subir plantilla', category='x',
        data_start_row=5, research_fields=[],
    )
    excel = _module(
        'excel_workflow',
        prepare_research=prepare_research or (lambda *a, **k: SimpleNamespace(schema=default_schema, researchable_count=1)),
        build_preview_rows=lambda *a, **k: [],
        write_validated_workbook=write_validated_workbook or (lambda *a, **k: None),
    )
    runner = _module(
        'research_runner',
        run_research_for_preview_once=run_research_for_preview_once or (lambda *a, **k: None),
    )
    legacy_stubs = {
        'excel_workflow': excel,
        'research_runner': runner,
        'price_workflow': _module(
            'price_workflow',
            prepare_price_research=lambda *a, **k: None,
            run_price_research_v21=lambda *a, **k: None,
            export_prices_xlsx=lambda *a, **k: None,
        ),
        'price_web_verifier': _module(
            'price_web_verifier', verify_price_offers_headless=lambda *a, **k: None,
        ),
        'image_workflow': _module(
            'image_workflow',
            prepare_image_research=lambda *a, **k: None,
            run_image_research=lambda *a, **k: None,
            download_image_records=lambda *a, **k: None,
        ),
        'video_workflow': _module(
            'video_workflow',
            prepare_video_research=lambda *a, **k: None,
            run_video_research=lambda *a, **k: None,
            download_video_records=lambda *a, **k: None,
        ),
    }

    import app
    sentinel = object()
    saved_attr = getattr(app, 'workflows', sentinel)
    saved_module = sys.modules.pop('app.workflows', None)
    try:
        with patch.dict(sys.modules, legacy_stubs, clear=False):
            import app.workflows as workflows
            yield workflows, excel
    finally:
        sys.modules.pop('app.workflows', None)
        if saved_module is not None:
            sys.modules['app.workflows'] = saved_module
        if saved_attr is sentinel:
            if hasattr(app, 'workflows'):
                delattr(app, 'workflows')
        else:
            app.workflows = saved_attr


class CharacteristicsWorkflowTests(unittest.IsolatedAsyncioTestCase):
    async def test_auto_mode_passes_resolved_part_number_to_legacy_and_returns_canonical_identity(self):
        captured = {}

        def prepare_research(path, identifier):
            captured['identifier'] = identifier
            schema = SimpleNamespace(
                family='falabella', sheet_name='Subir plantilla', category='x',
                data_start_row=5, research_fields=[],
            )
            return SimpleNamespace(schema=schema, researchable_count=1)

        async def run_once(*args, **kwargs):
            raw = {
                'producto': {'marca':'EPSON','modelo':'L3350','confidence':98},
                'especificaciones_completas': [{
                    'key':'resolution','label':'Resolución','value':'5760 x 1440','unit':'dpi',
                    'status':'CONFIRMED','confidence':97,'source_url':'https://epson.test/spec.pdf',
                    'source_type':'OFFICIAL_PDF','applies_to':'C11CL62301'
                }]
            }
            validation = SimpleNamespace(raw=raw, accepted=[1], rejected=[])
            return SimpleNamespace(validation=validation, raw_paths=[], followup_performed=False)

        with isolated_workflows(
            prepare_research=prepare_research,
            run_research_for_preview_once=run_once,
        ) as (workflows, _):
            with tempfile.TemporaryDirectory() as td:
                path = Path(td)/'template.xlsx'
                make_workbook(path)
                job = SimpleNamespace(id='job1', directory=Path(td), payload={})
                with patch.object(workflows, 'chatgpt_session', return_value=FakeSession()):
                    result = await workflows.run_characteristics(job, '', path, lambda *a, **k: None)

            self.assertEqual(captured['identifier'], 'C11CL62301')
            self.assertEqual(result['input_mode'], 'auto')
            self.assertEqual(result['detected_identifier'], 'C11CL62301')
            self.assertEqual(result['identity']['manufacturer_part_number'], 'C11CL62301')
            self.assertEqual(result['identity']['commercial_model'], 'L3350')
            self.assertTrue(result['qa_ready'])
            self.assertEqual(job.payload['canonical_identity'].manufacturer_part_number, 'C11CL62301')


class CharacteristicsExcelArtifactTests(unittest.TestCase):
    def test_generate_excel_finalizes_and_marks_only_qa_passing_artifact_completed(self):
        from app.product_evidence import MasterSpecification
        from app.product_identity import CanonicalIdentity

        with tempfile.TemporaryDirectory() as td:
            output = Path(td) / 'legacy.xlsx'

            def writer(*args, **kwargs):
                make_workbook(output)
                return output

            with isolated_workflows(write_validated_workbook=writer) as (workflows, _):
                identity = CanonicalIdentity(
                    brand='EPSON', manufacturer_part_number='C11CL62301', commercial_model='L3350', confidence=99,
                    sources=[{'url':'https://epson.test','source_type':'OFFICIAL_PRODUCT'}],
                )
                job = SimpleNamespace(directory=Path(td), payload={
                    'preparation': object(), 'validation': object(), 'canonical_identity': identity,
                    'master_specifications': [MasterSpecification(
                        'resolution','Resolución','5760 x 1440','dpi','CONFIRMED',97,
                        'https://epson.test/spec.pdf','OFFICIAL_PDF','Ficha',2,'5760 x 1440','C11CL62301'
                    )],
                })
                result_path = workflows.generate_excel(job)

            self.assertIn('COMPLETADO', result_path.stem)
            wb = load_workbook(result_path)
            ws = wb['Subir plantilla']
            self.assertEqual(ws['A2'].value, 'EPSON')
            self.assertEqual(ws['B2'].value, 'C11CL62301')
            self.assertEqual(ws['C2'].value, 'L3350')
            self.assertIn('ESPECIFICACIONES_COMPLETAS', wb.sheetnames)
            wb.close()

    def test_generate_excel_renames_failed_legacy_completed_artifact_to_no_validado(self):
        from app.product_identity import CanonicalIdentity

        with tempfile.TemporaryDirectory() as td:
            output = Path(td) / 'legacy_COMPLETADO.xlsx'

            def writer(*args, **kwargs):
                make_workbook(output)
                return output

            with isolated_workflows(write_validated_workbook=writer) as (workflows, _):
                job = SimpleNamespace(directory=Path(td), payload={
                    'preparation': object(), 'validation': object(),
                    'canonical_identity': CanonicalIdentity(brand='EPSON', manufacturer_part_number='C11CL62301'),
                    'master_specifications': [],
                })
                with self.assertRaisesRegex(ValueError, 'PRODUCT_WORKBOOK_QA_FAILED'):
                    workflows.generate_excel(job)

            names = [p.name for p in Path(td).glob('*.xlsx')]
            self.assertTrue(any('NO_VALIDADO' in name for name in names), names)
            self.assertFalse(any('COMPLETADO' in name for name in names), names)

    def test_legacy_stubs_do_not_leak_to_other_tests(self):
        sentinel_excel = sys.modules.get('excel_workflow')
        sentinel_runner = sys.modules.get('research_runner')
        with isolated_workflows():
            self.assertIsNot(sys.modules.get('excel_workflow'), sentinel_excel)
            self.assertIsNot(sys.modules.get('research_runner'), sentinel_runner)
        self.assertIs(sys.modules.get('excel_workflow'), sentinel_excel)
        self.assertIs(sys.modules.get('research_runner'), sentinel_runner)


if __name__ == '__main__':
    unittest.main()

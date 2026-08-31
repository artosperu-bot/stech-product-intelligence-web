import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from app.marketplace_template import analyze_marketplace_template, choose_manual_slot, iter_required_fields


def make_falabella(path: Path, with_evidence=False):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Subir plantilla'
    ws.append(['Principales'] * 6)
    ws.append([
        '- Value: Esto es un párrafo', '- Value: JBLDEMO123', '- Value: ABC-1000-202',
        '- Value: 1234567890123', '- Value: 888 - Impresoras', '- Value: Nuevo',
    ])
    ws.append([' ', ' ', ' ', '( Optional )', ' ', ' '])
    ws.append(['Nombre #39', 'SKU del vendedor #29', 'Modelo #32', 'Código de barras #56', 'Categoría primaria #1', 'Condición del Producto #22'])
    ws.append(['', '', 'C11CL62301', '', '', ''])
    ws.append(['', '', 'C11CL65301', '', '', ''])
    ws.append(['Esto es un párrafo', 'JBLDEMO123', 'ABC-1000-202', '1234567890123', '888 - Impresoras', 'Nuevo'])

    cats = wb.create_sheet('Categorías')
    cats.append(['Categorías'])
    cats.append(['888 - Impresoras'])
    opts = wb.create_sheet('Opciones')
    opts.append(['ConditionType'])
    opts.append(['Nuevo'])
    opts.append(['Open Box'])
    if with_evidence:
        ev = wb.create_sheet('IA_EVIDENCIA')
        ev.append(['Part Number', 'Estado', 'Confianza'])
        ev.append(['OLD-WRONG-999', 'CONFIRMADO', 100])
    wb.save(path)


def make_ripley(path: Path):
    wb = Workbook()
    data = wb.active
    data.title = 'Data'
    data.append(['Categoría', 'Nombre', 'EAN/UPC', 'Modelo', 'SKU Seller', 'Estado', 'Precio'])
    data.append(['categoria', 'nombre', 'ean', 'modelo', 'sku_seller', 'state', 'price'])
    data.append(['CAT-MOUSE', 'Mouse A', '1234567890123', 'RGBM-02-BK', 'SELLER-A', 'Nuevo', 49.9])
    data.append(['CAT-MOUSE', 'Mouse B', '1234567890124', 'RGBM-02-WH', 'SELLER-B', 'Nuevo', 49.9])

    cols = wb.create_sheet('Columns')
    cols.append(['Código', 'Etiqueta', 'Descripción', 'Valor de ejemplo', 'CAT-MOUSE'])
    cols.append(['categoria', 'Categoría', '', '', 'REQUIRED'])
    cols.append(['nombre', 'Nombre', '', 'Nombre de ejemplo', 'REQUIRED'])
    cols.append(['ean', 'EAN/UPC', '', '1234567890123', 'RECOMMENDED'])
    cols.append(['modelo', 'Modelo', '', 'ABC-100', 'REQUIRED'])
    cols.append(['sku_seller', 'SKU Seller', '', 'SELLER-1', 'REQUIRED'])
    cols.append(['state', 'Estado', '', '', 'REQUIRED'])
    cols.append(['price', 'Precio', '', '', 'REQUIRED'])

    ref = wb.create_sheet('ReferenceData')
    ref.append(['state'])
    ref.append(['Nuevo'])
    ref.append(['Usado'])
    wb.save(path)


class MarketplaceTemplateTests(unittest.TestCase):
    def test_falabella_detects_all_real_rows_and_ignores_examples_and_old_evidence(self):
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / 'falabella.xlsx'
            make_falabella(path, with_evidence=True)
            profile = analyze_marketplace_template(path)

        self.assertEqual(profile.marketplace, 'falabella')
        self.assertEqual(profile.sheet_name, 'Subir plantilla')
        self.assertEqual([(p.row, p.identifier, p.identifier_type) for p in profile.products], [
            (5, 'C11CL62301', 'PART_NUMBER'),
            (6, 'C11CL65301', 'PART_NUMBER'),
        ])
        self.assertNotIn('OLD-WRONG-999', [p.identifier for p in profile.products])
        self.assertEqual(profile.category_options, ('888 - Impresoras',))
        condition = profile.field_by_label('Condición del Producto')
        self.assertEqual(condition.valid_values, ('Nuevo', 'Open Box'))

    def test_manual_identifier_resolves_matching_row(self):
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / 'falabella.xlsx'
            make_falabella(path)
            profile = analyze_marketplace_template(path)
        slot = choose_manual_slot(profile, 'C11CL65301')
        self.assertIsNotNone(slot)
        self.assertEqual(slot.row, 6)

    def test_ripley_detects_dual_headers_products_requirements_and_reference_values(self):
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / 'ripley.xlsx'
            make_ripley(path)
            profile = analyze_marketplace_template(path)

        self.assertEqual(profile.marketplace, 'ripley')
        self.assertEqual(profile.header_rows, (1, 2))
        self.assertEqual([(p.row, p.identifier) for p in profile.products], [(3, 'RGBM-02-BK'), (4, 'RGBM-02-WH')])
        required_codes = {f.code for f in iter_required_fields(profile, profile.products[0])}
        self.assertEqual(required_codes, {'categoria', 'nombre', 'modelo', 'sku_seller', 'state', 'price'})
        self.assertEqual(profile.field_by_code('state').valid_values, ('Nuevo', 'Usado'))


if __name__ == '__main__':
    unittest.main()

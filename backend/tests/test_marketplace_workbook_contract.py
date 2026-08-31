import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.marketplace_template import analyze_marketplace_template
from app.marketplace_workbook import ProductWriteRecord, write_marketplace_workbook
from app.product_identity import CanonicalIdentity


def _identity():
    return CanonicalIdentity(
        brand='JBL',
        manufacturer_part_number='JBLTEST100',
        commercial_model='Test 100',
        ean_upc_gtin=[],
        confidence=99,
        sources=[{'url': 'https://official.example/test', 'source_type': 'OFFICIAL_PRODUCT'}],
    )


def _make_book(path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Subir plantilla'
    ws.append(['Principales'] * 6)
    ws.append([
        'Máximo 60 caracteres.',
        'Marca oficial.',
        'Modelo comercial.',
        'Descripción comercial. Máximo 40 caracteres.',
        'Selecciona color permitido.',
        'MPN del fabricante.',
    ])
    ws.append([' ', ' ', ' ', ' ', ' ', ' '])
    ws.append(['Nombre #39', 'Marca #26', 'Modelo #32', 'Descripción #53', 'Color #100', 'SKU del vendedor #29'])
    ws.append(['', '', 'JBLTEST100', '', '', ''])

    opts = wb.create_sheet('Opciones')
    opts.append(['Color'])
    opts.append(['Negro'])
    opts.append(['Azul'])
    wb.save(path)


class MarketplaceWorkbookContractTests(unittest.TestCase):
    def test_writer_rejects_out_of_catalog_option_and_overlong_text(self):
        with tempfile.TemporaryDirectory() as td:
            src = Path(td) / 'source.xlsx'
            out = Path(td) / 'result.xlsx'
            _make_book(src)
            profile = analyze_marketplace_template(src)
            record = ProductWriteRecord(
                slot=profile.products[0],
                identity=_identity(),
                preview_rows=[
                    {'field': 'Nombre #39', 'value': 'JBL Test 100', 'status': 'CONFIRMED', 'confidence': 99},
                    {'field': 'Descripción #53', 'value': 'X' * 41, 'status': 'CONFIRMED', 'confidence': 99},
                    {'field': 'Color #100', 'value': 'Verde', 'status': 'CONFIRMED', 'confidence': 99},
                ],
            )
            qa = write_marketplace_workbook(src, out, profile, [record])
            self.assertFalse(qa.ok)
            wb = load_workbook(out)
            ws = wb['Subir plantilla']
            self.assertIsNone(ws['D5'].value)
            self.assertIsNone(ws['E5'].value)

            evidence = wb['IA_EVIDENCIA']
            rejected = {
                row[1].value: (row[3].value, row[4].value, row[10].value)
                for row in evidence.iter_rows(min_row=2)
                if row[1].value in {'Descripción #53', 'Color #100'}
            }
            self.assertEqual(rejected['Descripción #53'][0], None)
            self.assertEqual(rejected['Descripción #53'][1], 'REJECTED_TEMPLATE')
            self.assertIn('MAX_CHARS_EXCEEDED:40', rejected['Descripción #53'][2])
            self.assertEqual(rejected['Color #100'][0], None)
            self.assertEqual(rejected['Color #100'][1], 'REJECTED_TEMPLATE')
            self.assertIn('VALUE_NOT_ALLOWED', rejected['Color #100'][2])
            wb.close()

            warnings = qa.products[0].warnings
            self.assertTrue(any('MAX_CHARS_EXCEEDED:40' in item for item in warnings))
            self.assertTrue(any('VALUE_NOT_ALLOWED' in item for item in warnings))


if __name__ == '__main__':
    unittest.main()

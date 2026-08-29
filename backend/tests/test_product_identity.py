import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from app.product_identity import (
    CanonicalIdentity,
    canonical_identity_from_raw,
    choose_research_identifier,
    enrich_identity_from_workbook_evidence,
    extract_identity_candidates,
)


def make_workbook(path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Subir plantilla'
    ws.append(['Marca #26', 'SKU del vendedor #29', 'Modelo #32', 'Código de barras #56', 'Nombre #39'])
    ws.append(['EPSON', '', 'C11CL62301', '0103439891745', 'Impresora Epson EcoTank L3350'])
    wb.save(path)


def make_workbook_with_leading_empty_cells(path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Subir plantilla'
    headers = ['Marca #26', 'SKU del vendedor #29', 'Modelo #32', 'Código de barras #56', 'Nombre #39']
    values = ['EPSON', '', 'C11CL62301', '0103439891745', 'Impresora Epson EcoTank L3350']
    for column, value in enumerate(headers, start=2):
        ws.cell(row=4, column=column, value=value)
    for column, value in enumerate(values, start=2):
        ws.cell(row=5, column=column, value=value)
    wb.save(path)


class ProductIdentityTests(unittest.TestCase):
    def test_extracts_part_number_from_model_and_barcode_separately(self):
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / 'template.xlsx'
            make_workbook(path)
            candidates = extract_identity_candidates(path)
        by_value = {c.value: c for c in candidates}
        self.assertEqual(by_value['C11CL62301'].kind, 'PART_NUMBER')
        self.assertEqual(by_value['C11CL62301'].field_name, 'Modelo #32')
        self.assertEqual(by_value['0103439891745'].kind, 'EAN_UPC_GTIN')

    def test_extracts_candidates_when_first_columns_and_rows_are_empty(self):
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / 'template.xlsx'
            make_workbook_with_leading_empty_cells(path)
            candidates = extract_identity_candidates(path)
        by_value = {c.value: c for c in candidates}
        self.assertEqual(by_value['C11CL62301'].row, 5)
        self.assertEqual(by_value['C11CL62301'].kind, 'PART_NUMBER')
        self.assertEqual(by_value['0103439891745'].row, 5)

    def test_manual_identifier_wins_and_auto_uses_best_candidate(self):
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / 'template.xlsx'
            make_workbook(path)
            candidates = extract_identity_candidates(path)
        self.assertEqual(
            choose_research_identifier(' C11CL62301 ', candidates),
            ('manual', 'C11CL62301', 'PART_NUMBER'),
        )
        self.assertEqual(
            choose_research_identifier('', candidates),
            ('auto', 'C11CL62301', 'PART_NUMBER'),
        )

    def test_canonical_identity_keeps_mpn_model_and_barcode_distinct(self):
        identity = canonical_identity_from_raw({
            'producto': {
                'marca': 'EPSON',
                'part_number': 'C11CL62301',
                'modelo': 'L3350',
                'ean': '0103439891745',
                'confidence': 98,
            }
        })
        self.assertIsInstance(identity, CanonicalIdentity)
        self.assertEqual(identity.manufacturer_part_number, 'C11CL62301')
        self.assertEqual(identity.commercial_model, 'L3350')
        self.assertEqual(identity.ean_upc_gtin, ['0103439891745'])

    def test_confirmed_evidence_part_number_outranks_stale_model_candidate(self):
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / 'template.xlsx'
            wb = Workbook()
            ws = wb.active
            ws.title = 'Subir plantilla'
            ws.append(['Marca #26', 'SKU del vendedor #29', 'Modelo #32', 'Código de barras #56', 'Nombre #39'])
            ws.append(['', '', 'jblq350wlblkam', '', ''])
            ev = wb.create_sheet('IA_EVIDENCIA')
            ev.append(['Part Number', 'Campo', 'Valor', 'Estado', 'Confianza', 'Fuente'])
            ev.append(['C11CL62301', 'Modelo #32', 'L3350', 'CONFIRMADO', 99, 'https://epson.test/support'])
            wb.save(path)
            candidates = extract_identity_candidates(path)
        self.assertEqual(
            choose_research_identifier('', candidates),
            ('auto', 'C11CL62301', 'PART_NUMBER'),
        )

    def test_workbook_evidence_fills_missing_brand_and_model_without_overwriting_mpn(self):
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / 'template.xlsx'
            wb = Workbook()
            ws = wb.active
            ws.title = 'Subir plantilla'
            ws.append(['Marca #26', 'SKU del vendedor #29', 'Modelo #32'])
            ws.append(['', '', 'stale-other-product'])
            ev = wb.create_sheet('IA_EVIDENCIA')
            ev.append(['Part Number', 'Campo', 'Valor', 'Estado', 'Confianza', 'Fuente'])
            ev.append(['C11CL62301', 'Marca #26', 'EPSON', 'CONFIRMADO', 100, 'https://epson.test'])
            ev.append(['C11CL62301', 'Modelo #32', 'L3350', 'CONFIRMADO', 99, 'https://epson.test'])
            wb.save(path)
            base = CanonicalIdentity(manufacturer_part_number='C11CL62301')
            enriched = enrich_identity_from_workbook_evidence(path, base)
        self.assertEqual(enriched.brand, 'EPSON')
        self.assertEqual(enriched.commercial_model, 'L3350')
        self.assertEqual(enriched.manufacturer_part_number, 'C11CL62301')

    def test_auto_raises_when_no_candidate_exists(self):
        with self.assertRaisesRegex(ValueError, 'IDENTITY_CANDIDATE_NOT_FOUND'):
            choose_research_identifier('', [])


if __name__ == '__main__':
    unittest.main()

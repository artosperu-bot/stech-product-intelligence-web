import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.product_evidence import MasterSpecification
from app.product_identity import CanonicalIdentity
from app.product_workbook import assert_product_workbook_qa, finalize_product_workbook


def make_bad_workbook(path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Subir plantilla'
    ws.append(['Principales'] * 8)
    ws.append(['Instructions'] * 8)
    ws.append(['89'] * 8)
    ws.append([
        'Nombre #39', 'Marca #26', 'Modelo #32', 'Descripción #53',
        'Categoría primaria #1', 'SKU del vendedor #29', 'Código de barras #56',
        'CalidadDeImpresion #1696',
    ])
    ws.append([None, None, 'jblq350wlblkam', None, None, None, None, 'Alta resolución'])
    ws.append([None, None, 'jblendurrun3btbam', None, None, None, None, None])

    ev = wb.create_sheet('IA_EVIDENCIA')
    ev.append(['Part Number','Campo','Valor','Estado','Confianza','Fuente','Tipo fuente','Evidencia','Observación'])
    ev.append(['C11CL62301','Modelo #32','L3350','CONFIRMADO',99,'https://epson.test/support','SOPORTE','Exact model','89'])
    wb.save(path)


class ProductWorkbookTests(unittest.TestCase):
    def setUp(self):
        self.identity = CanonicalIdentity(
            brand='EPSON', manufacturer_part_number='C11CL62301', commercial_model='L3350',
            ean_upc_gtin=['0103439891745'], confidence=99,
            sources=[{'url':'https://epson.test/product','source_type':'OFFICIAL_PRODUCT'}],
        )
        self.specs = [
            MasterSpecification(
                'resolution', 'Resolución máxima', '5760 x 1440', 'dpi', 'CONFIRMED', 97,
                'https://epson.test/spec.pdf', 'OFFICIAL_PDF', 'Ficha técnica', 2,
                '5760 x 1440 dpi', 'C11CL62301',
            )
        ]

    def test_normalizes_identity_and_purges_unrelated_rows(self):
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / 'bad.xlsx'
            make_bad_workbook(path)
            qa = finalize_product_workbook(path, self.identity, self.specs, [])
            wb = load_workbook(path, data_only=False)
            ws = wb['Subir plantilla']
            self.assertEqual(ws['B5'].value, 'EPSON')
            self.assertEqual(ws['C5'].value, 'L3350')
            self.assertEqual(ws['F5'].value, 'C11CL62301')
            self.assertEqual(ws['G5'].value, '0103439891745')
            self.assertTrue(all(ws.cell(6, c).value is None for c in range(1, 9)))
            wb.close()
            self.assertTrue(qa.ok, qa.errors)

    def test_creates_master_and_strengthened_evidence_sheets(self):
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / 'bad.xlsx'
            make_bad_workbook(path)
            finalize_product_workbook(path, self.identity, self.specs, [])
            wb = load_workbook(path, data_only=False)
            self.assertIn('ESPECIFICACIONES_COMPLETAS', wb.sheetnames)
            master = wb['ESPECIFICACIONES_COMPLETAS']
            self.assertEqual(master['A1'].value, 'Part Number')
            self.assertEqual(master['C2'].value, 'Resolución máxima')
            evidence = wb['IA_EVIDENCIA']
            headers = [cell.value for cell in evidence[1]]
            self.assertIn('Página PDF', headers)
            self.assertIn('Valor escrito', headers)
            rows = list(evidence.iter_rows(min_row=2, values_only=True))
            self.assertTrue(any(row[1] == 'SKU del vendedor #29' and row[3] == 'C11CL62301' for row in rows))
            self.assertFalse(any(row[3] == '89' and row[4] == 'CONFIRMED' for row in rows))
            wb.close()

    def test_master_sheet_falls_back_to_confirmed_legacy_evidence_when_no_extension_specs(self):
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / 'bad.xlsx'
            make_bad_workbook(path)
            qa = finalize_product_workbook(path, self.identity, [], [])
            self.assertTrue(qa.ok, qa.errors)
            wb = load_workbook(path, data_only=False)
            rows = list(wb['ESPECIFICACIONES_COMPLETAS'].iter_rows(min_row=2, values_only=True))
            self.assertTrue(any(row[2] == 'Modelo #32' and row[3] == 'L3350' for row in rows))
            wb.close()

    def test_qa_blocks_missing_critical_identity(self):
        bad_identity = CanonicalIdentity(brand='EPSON', manufacturer_part_number='', commercial_model='L3350')
        with tempfile.TemporaryDirectory() as td:
            path = Path(td) / 'bad.xlsx'
            make_bad_workbook(path)
            qa = finalize_product_workbook(path, bad_identity, [], [])
            self.assertFalse(qa.ok)
            with self.assertRaisesRegex(ValueError, 'PRODUCT_WORKBOOK_QA_FAILED'):
                assert_product_workbook_qa(qa)


if __name__ == '__main__':
    unittest.main()
